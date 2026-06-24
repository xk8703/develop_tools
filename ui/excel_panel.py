"""
Excel 处理面板
"""
from copy import copy
from pathlib import Path
from tkinter import filedialog

import customtkinter as ctk
from openpyxl import load_workbook


class ExcelPanel(ctk.CTkFrame):
    """Excel 处理工具面板"""

    def __init__(self, master):
        super().__init__(master)
        self.pack(fill="both", expand=True)

        self.file_path_var = ctk.StringVar(value="")
        self.output_dir_var = ctk.StringVar(value="")
        self.header_row_var = ctk.StringVar(value="3")

        self._create_ui()

    def _create_ui(self):
        frame = ctk.CTkFrame(self)
        frame.pack(fill="both", expand=True, padx=20, pady=20)

        title = ctk.CTkLabel(
            frame,
            text="Excel 处理",
            font=ctk.CTkFont(size=18, weight="bold")
        )
        title.pack(pady=(0, 20))

        file_row = ctk.CTkFrame(frame)
        file_row.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(file_row, text="Excel文件:", width=90, anchor="w").pack(
            side="left", padx=(10, 5)
        )
        ctk.CTkEntry(file_row, textvariable=self.file_path_var).pack(
            side="left", fill="x", expand=True, padx=(0, 5)
        )
        ctk.CTkButton(file_row, text="浏览", width=70, command=self._browse_file).pack(
            side="left", padx=(0, 10)
        )

        output_row = ctk.CTkFrame(frame)
        output_row.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(output_row, text="输出目录:", width=90, anchor="w").pack(
            side="left", padx=(10, 5)
        )
        ctk.CTkEntry(output_row, textvariable=self.output_dir_var).pack(
            side="left", fill="x", expand=True, padx=(0, 5)
        )
        ctk.CTkButton(output_row, text="浏览", width=70, command=self._browse_output_dir).pack(
            side="left", padx=(0, 10)
        )

        option_row = ctk.CTkFrame(frame)
        option_row.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(option_row, text="表头行号:", width=90, anchor="w").pack(
            side="left", padx=(10, 5)
        )
        ctk.CTkEntry(option_row, textvariable=self.header_row_var, width=80).pack(
            side="left", padx=(0, 10)
        )

        action_row = ctk.CTkFrame(frame, fg_color="transparent")
        action_row.pack(fill="x", pady=(0, 10))

        ctk.CTkButton(
            action_row,
            text="单据号取消合并",
            width=140,
            command=self._unmerge_receipt_no
        ).pack(side="left", padx=(10, 5))

        self.result_text = ctk.CTkTextbox(frame, height=180)
        self.result_text.pack(fill="both", expand=True, padx=10, pady=(10, 0))
        self._set_result("请选择 Excel 文件")

    def _browse_file(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        if file_path:
            self.file_path_var.set(file_path)
            if not self.output_dir_var.get():
                self.output_dir_var.set(str(Path(file_path).parent))

    def _browse_output_dir(self):
        output_dir = filedialog.askdirectory()
        if output_dir:
            self.output_dir_var.set(output_dir)

    def _unmerge_receipt_no(self):
        source_text = self.file_path_var.get().strip()
        if not source_text:
            self._set_result("请选择 Excel 文件")
            return

        source = Path(source_text)
        if not source.exists():
            self._set_result("Excel 文件不存在")
            return

        output_dir_text = self.output_dir_var.get().strip()
        output_dir = Path(output_dir_text) if output_dir_text else source.parent
        output = output_dir / f"{source.stem}_单据号取消合并{source.suffix}"

        try:
            header_row = int(self.header_row_var.get().strip())
            if header_row < 1:
                raise ValueError
        except ValueError:
            self._set_result("表头行号必须是大于 0 的整数")
            return

        try:
            sheet_name, target_col, unmerged_count = self._process_receipt_no(source, output, header_row)
        except Exception as e:
            self._set_result(f"处理失败: {e}")
            return

        self._set_result(
            f"处理完成\n"
            f"工作表: {sheet_name}\n"
            f"表头行号: {header_row}\n"
            f"单据号列: {target_col}\n"
            f"取消合并数量: {unmerged_count}\n"
            f"输出文件: {output}"
        )

    def _process_receipt_no(self, source, output, header_row):
        wb = load_workbook(source)
        ws = wb.active
        data_start_row = header_row + 1

        target_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(header_row, col).value == "单据号":
                target_col = col
                break

        if target_col is None:
            raise ValueError(f"第 {header_row} 行未找到表头：单据号")

        target_ranges = [
            merge_range
            for merge_range in list(ws.merged_cells.ranges)
            if merge_range.min_col == target_col
            and merge_range.max_col == target_col
            and merge_range.max_row >= data_start_row
        ]

        for merge_range in target_ranges:
            start_row = max(merge_range.min_row, data_start_row)
            value = ws.cell(merge_range.min_row, target_col).value
            style_source = ws.cell(merge_range.min_row, target_col)
            fill_style = {
                "font": copy(style_source.font),
                "fill": copy(style_source.fill),
                "border": copy(style_source.border),
                "alignment": copy(style_source.alignment),
                "number_format": style_source.number_format,
                "protection": copy(style_source.protection),
            }
            ws.unmerge_cells(str(merge_range))
            for row in range(start_row, merge_range.max_row + 1):
                cell = ws.cell(row, target_col)
                cell.value = value
                cell.font = fill_style["font"]
                cell.fill = fill_style["fill"]
                cell.border = fill_style["border"]
                cell.alignment = fill_style["alignment"]
                cell.number_format = fill_style["number_format"]
                cell.protection = fill_style["protection"]

        output.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output)
        return ws.title, target_col, len(target_ranges)

    def _set_result(self, text):
        self.result_text.configure(state="normal")
        self.result_text.delete("1.0", "end")
        self.result_text.insert("1.0", text)
        self.result_text.configure(state="disabled")
